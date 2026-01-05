from fastapi import FastAPI, APIRouter
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class FormSubmission(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    mobile_no: str
    shop_name: str
    owner_name: str
    ind_name: str
    area_pin_code: str
    address: str
    city: str
    dist: str
    state: str
    country: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class FormSubmissionCreate(BaseModel):
    mobile_no: str = Field(..., min_length=10, max_length=10, pattern=r"^\d{10}$")
    shop_name: str = Field(..., min_length=1)
    owner_name: str = Field(..., min_length=1)
    ind_name: str = Field(..., min_length=1)
    area_pin_code: str = Field(..., min_length=6, max_length=6, pattern=r"^\d{6}$")
    address: str = Field(..., min_length=1)
    city: str = Field(..., min_length=1)
    dist: str = Field(..., min_length=1)
    state: str = Field(..., min_length=1)
    country: str = Field(..., min_length=1)


class FormSubmissionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str
    mobile_no: str
    shop_name: str
    owner_name: str
    ind_name: str
    area_pin_code: str
    address: str
    city: str
    dist: str
    state: str
    country: str
    created_at: str


# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Form API is running"}


@api_router.post("/submissions", response_model=FormSubmissionResponse)
async def create_submission(input: FormSubmissionCreate):
    submission_dict = input.model_dump()
    submission_obj = FormSubmission(**submission_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = submission_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    _ = await db.form_submissions.insert_one(doc)
    
    return FormSubmissionResponse(
        id=doc['id'],
        mobile_no=doc['mobile_no'],
        shop_name=doc['shop_name'],
        owner_name=doc['owner_name'],
        ind_name=doc['ind_name'],
        area_pin_code=doc['area_pin_code'],
        address=doc['address'],
        city=doc['city'],
        dist=doc['dist'],
        state=doc['state'],
        country=doc['country'],
        created_at=doc['created_at']
    )


@api_router.get("/submissions", response_model=List[FormSubmissionResponse])
async def get_submissions():
    # Exclude MongoDB's _id field from the query results
    submissions = await db.form_submissions.find({}, {"_id": 0}).to_list(1000)
    
    result = []
    for sub in submissions:
        result.append(FormSubmissionResponse(
            id=sub['id'],
            mobile_no=sub['mobile_no'],
            shop_name=sub['shop_name'],
            owner_name=sub['owner_name'],
            ind_name=sub['ind_name'],
            area_pin_code=sub['area_pin_code'],
            address=sub['address'],
            city=sub['city'],
            dist=sub['dist'],
            state=sub['state'],
            country=sub['country'],
            created_at=sub['created_at']
        ))
    
    return result


@api_router.get("/submissions/export")
async def export_submissions_excel():
    """Export all submissions to Excel file"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Fetch all submissions
    submissions = await db.form_submissions.find({}, {"_id": 0}).to_list(1000)
    
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Submissions"
    
    # Define headers
    headers = [
        "S.No", "Mobile No.", "Shop Name", "Owner Name", "Industry Name",
        "Pin Code", "Address", "City", "District", "State", "Country", "Submitted At"
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, sub in enumerate(submissions, 2):
        row_data = [
            row_idx - 1,
            sub.get('mobile_no', ''),
            sub.get('shop_name', ''),
            sub.get('owner_name', ''),
            sub.get('ind_name', ''),
            sub.get('area_pin_code', ''),
            sub.get('address', ''),
            sub.get('city', ''),
            sub.get('dist', ''),
            sub.get('state', ''),
            sub.get('country', ''),
            sub.get('created_at', '')
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [8, 15, 25, 20, 20, 12, 40, 15, 15, 15, 15, 22]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"form_submissions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.delete("/submissions/{submission_id}")
async def delete_submission(submission_id: str):
    """Delete a submission by ID"""
    result = await db.form_submissions.delete_one({"id": submission_id})
    if result.deleted_count == 0:
        return {"success": False, "message": "Submission not found"}
    return {"success": True, "message": "Submission deleted"}


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
